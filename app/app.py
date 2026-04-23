import re
import configparser
from bs4 import BeautifulSoup
import requests
from requests.exceptions import RequestException
from .config import get_config_value, get_pdf_path, set_config_value, get_config_path
from .data_list import sites
import app.utils_scrapping as scrapping
from .utils_pdf import download_pdf, delete_pdfs
import datetime
from datetime import timedelta
from datetime import date
from datetime import datetime
from openpyxl import load_workbook, Workbook
import sys
import os
import subprocess
from io import StringIO
from ressources.colors import bg_color, bg_color_light, bg_color, text_light, text_medium, text_dark
from app.utils_format import check_and_return_value
import threading
import ssl
import locale
from dateutil.easter import easter

from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QVBoxLayout, QWidget, QPushButton, QFileDialog, QTextEdit, QLineEdit
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QInputDialog, QProgressBar
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import QTimer
from PyQt5 import QtWidgets, QtCore

# Chemin absolu vers config.ini basé sur l'emplacement de votre script
config_path = get_config_path()
print(config_path)
config = configparser.ConfigParser()
config.read(config_path)
locale.setlocale(locale.LC_TIME, 'fr_FR')
now = datetime.now().date()
yesterday = now - timedelta(days=1)

# Ajd 'vendredi'
day_of_week = now.strftime("%A")
date_actuelle = datetime.now()
jour_actuel = date_actuelle.weekday()

# Hier 'd/m/Y
date_yesterday = yesterday.strftime("%d/%m/%Y")

# Hier 'jeudi'
yesterday_day_of_week = yesterday.strftime("%A")

# Hier 'jeudi 01 juin 2023
yesterday_holiday = yesterday.strftime("%A %d %B")

correspondance_jours = {0: 4, 1: 0, 2: 1, 3: 2, 4: 3}
jour_attendu = correspondance_jours.get(jour_actuel)

def get_uk_holidays(year):
    # Jours fériés fixes
    holidays_uk = [
        date(year, 1, 1),   # Jour de l'an
        date(year, 12, 25), # Noël
        date(year, 12, 26), # Lendemain de Noël
    ]
    
    # Premier lundi de mai
    may_day = date(year, 5, 1)
    while may_day.weekday() != 0:
        may_day += timedelta(days=1)
    holidays_uk.append(may_day)
    
    # Dernier lundi de mai
    spring_bank_holiday = date(year, 5, 31)
    while spring_bank_holiday.weekday() != 0:
        spring_bank_holiday -= timedelta(days=1)
    holidays_uk.append(spring_bank_holiday)
    
    # Dernier lundi d'août
    summer_bank_holiday = date(year, 8, 31)
    while summer_bank_holiday.weekday() != 0:
        summer_bank_holiday -= timedelta(days=1)
    holidays_uk.append(summer_bank_holiday)
    
    # Jours fériés variables basés sur Pâques
    good_friday = easter(year) - timedelta(days=2)
    holidays_uk.append(good_friday)
    
    easter_monday = easter(year) + timedelta(days=1)
    holidays_uk.append(easter_monday)

    holidays_uk_formatted = [
        holiday.strftime('%A %d %B').lower() for holiday in holidays_uk
    ]
    
    return holidays_uk_formatted

def get_french_holidays(year):
    # Jours fériés fixes
    holidays_french = [
        date(year, 1, 1),   # Jour de l'an
        date(year, 5, 1),   # Fête du travail
        date(year, 5, 8),   # Victoire des alliés
        date(year, 7, 14),  # Fête nationale
        date(year, 8, 15),  # Assomption
        date(year, 11, 1),  # Toussaint
        date(year, 11, 11), # Armistice
        date(year, 12, 25),# Noël
    ]
    
    # Jours fériés variables
    lundi_paques = easter(year) + timedelta(days=1)
    holidays_french.append(lundi_paques)
    
    ascension = easter(year) + timedelta(days=39)
    holidays_french.append(ascension)
    
    pentecote = easter(year) + timedelta(days=50)
    holidays_french.append(pentecote)

    # vendredi_saint = easter(year) - timedelta(days=2)
    # holidays_french.append(vendredi_saint)

    holidays_french_formatted = [
        holiday.strftime('%A %d %B').lower() for holiday in holidays_french
    ]

    return holidays_french_formatted

# Lire le chemin du fichier à partir du fichier config.ini
config = configparser.ConfigParser()
if os.path.exists(config_path):
    config.read(config_path)
    default_path_excel = config.get('SETTINGS', 'excel_path', fallback="")
    default_path_pdf = config.get('SETTINGS', 'pdf_path', fallback="")
    default_path_pdf_name = config.get('SETTINGS', 'name_pdf', fallback="")
else:
    default_path_excel = ""
    default_path_pdf = ""
    default_path_pdf_name = ""

class LoadingWindow(QMainWindow):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Chargement...")
        layout = QVBoxLayout()
        self.label = QLabel("Le script s'exécute, veuillez patienter...")
        layout.addWidget(self.label)
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def close(self):
        self.destroy()

class ScrappingSelectionDialog(QtWidgets.QDialog):
    """
    Une boîte de dialogue permettant à l'utilisateur de sélectionner les fonctions de scrapping à exécuter.
    
    Cette classe hérite de QDialog et est utilisée pour afficher une liste de fonctions de scrapping
    sous forme de cases à cocher. L'utilisateur peut sélectionner ou désélectionner les fonctions
    qu'il souhaite exécuter.
    
    Attributes:
    checkboxes (list): Une liste de widgets QCheckBox, chacun correspondant à une fonction de scrapping.
    toggle_select_button (QPushButton): Un bouton pour sélectionner ou désélectionner toutes les cases à cocher.
    ok_button (QPushButton): Un bouton pour fermer la boîte de dialogue et appliquer les sélections.
    
    Methods:
    __init__(self, scrapping_functions, selected_functions=[], parent=None): Initialise la boîte de dialogue.
    
    """
    def __init__(self, scrapping_functions, selected_functions=[], parent=None):
        """
        Initialise la boîte de dialogue avec une liste de fonctions de scrapping et une liste de fonctions sélectionnées.
        
        Parameters:
        scrapping_functions (list): Une liste de noms de fonctions de scrapping à afficher.
        selected_functions (list, optional): Une liste de noms de fonctions qui doivent être pré-sélectionnés. Par défaut, aucune fonction n'est sélectionnée.
        parent (QWidget, optional): Le widget parent de cette boîte de dialogue. Par défaut, il n'y a pas de parent.
        
        """
        super().__init__(parent)
        self.setWindowTitle('Sélectionner les Rates à récupérer.')
        
        self.layout = QtWidgets.QVBoxLayout(self)
        
        self.checkboxes = []  # Liste pour stocker les checkboxes

        # Bouton pour sélectionner/désélectionner toutes les cases à cocher
        self.toggle_select_button = QtWidgets.QPushButton('Sélectionner tous', self)
        self.toggle_select_button.clicked.connect(self.toggle_select_all)
        self.layout.addWidget(self.toggle_select_button)

        for func in scrapping_functions:
            checkbox = QtWidgets.QCheckBox(func)
            checkbox.setWhatsThis("Ceci est une case à cocher. Cochez ou décochez pour récupérer les cours de cette Rate.")
            if func in selected_functions:
                checkbox.setChecked(True)
            self.layout.addWidget(checkbox)
            self.checkboxes.append(checkbox)
        
        
        self.ok_button = QtWidgets.QPushButton('OK', self)
        self.ok_button.setWhatsThis("Cliquez sur OK pour valider les Rates à récuperer.")
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

        # Taille de la fenêtre
        self.setFixedSize(500, 700)

    def toggle_select_all(self):
        # Méthode pour basculer entre sélectionner et désélectionner toutes les cases à cocher
        if self.toggle_select_button.text() == 'Sélectionner tous':
            for checkbox in self.checkboxes:
                checkbox.setChecked(True)
            self.toggle_select_button.setText('Désélectionner tous')
        else:
            for checkbox in self.checkboxes:
                checkbox.setChecked(False)
            self.toggle_select_button.setText('Sélectionner tous')

    def get_selected_functions(self):
        selected_functions = []
        for checkbox in self.checkboxes:
            if checkbox.isChecked():
                selected_functions.append(checkbox.text())
        return selected_functions
    
    def count_checked_functions(self):
        """Compte le nombre de fonctions sélectionnées (checkboxes cochées)"""
        
        count = 0
        for checkbox in self.checkboxes:
            if checkbox.isChecked():
                count += 1
        return count


class MyApp(QtWidgets.QWidget):
    """
    Une application Qt qui permet de gérer et d'exécuter des scripts de scrapping.

    Cette classe hérite de QWidget et est utilisée pour créer l'interface utilisateur principale
    de l'application, gérer les interactions utilisateur, et exécuter les scripts de scrapping
    en fonction des paramètres sélectionnés par l'utilisateur.

    Attributes:
    config (ConfigParser): Un objet pour lire et écrire des configurations à un fichier INI.
    selected_scrapping_functions (list): Une liste des fonctions de scrapping sélectionnées par l'utilisateur.
    """
    def open_scrapping_selection_dialog(self):
        """
        Ouvre une boîte de dialogue qui permet à l'utilisateur de sélectionner les fonctions de scrapping à exécuter.
        """
        
        scrapping_functions = [site['func'] for site in sites]  # Créer une liste des noms de fonctions
        dialog = ScrappingSelectionDialog(scrapping_functions, self.selected_scrapping_functions, self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.selected_scrapping_functions = dialog.get_selected_functions()
            self.save_selected_functions()

    def save_selected_functions(self):
        """
        Sauvegarde les fonctions de scrapping sélectionnées par l'utilisateur dans un fichier de configuration.
        """
        
        selected_functions_str = ",".join(self.selected_scrapping_functions)
        self.config.set('SETTINGS', 'selected_functions', selected_functions_str)
        with open(config_path, 'w') as configfile:
            self.config.write(configfile)

    def load_selected_functions(self):
        """
        Charge les fonctions de scrapping sélectionnées depuis un fichier de configuration.
        """
        
        selected_functions_str = self.config.get('SETTINGS', 'selected_functions', fallback="")
        self.selected_scrapping_functions = selected_functions_str.split(",") if selected_functions_str else []

    def __init__(self):
        """
        Initialise l'application avec l'interface utilisateur et charge les configurations.
        """
        
        super().__init__()
        self.setWindowTitle("Cours des métaux")
        self.setGeometry(100, 100, 1000, 850)
        self.initUI()

        self.layout = QtWidgets.QVBoxLayout(self)

        self.select_scrapping_functions_button = QtWidgets.QPushButton('Sélectionner les fonctions de scrapping', self)
        self.select_scrapping_functions_button.clicked.connect(self.open_scrapping_selection_dialog)
        self.layout.addWidget(self.select_scrapping_functions_button)

        self.config = configparser.ConfigParser()
        self.config.read(config_path)
        self.load_selected_functions()
        auto_start = self.config.getboolean('SETTINGS', 'auto_start', fallback=False)
        self.param1_checkbox.setChecked(auto_start)

        if auto_start:
            if not self.path_pdf.text():
                QMessageBox.warning(
                self, "Chemin PDF manquant",
                "Le chemin d'accès au PDF est manquant. Veuillez le configurer avant de lancer le script.")
            else:    
                self.lancer_script(sites)
                QtCore.QTimer.singleShot(120000, self.close)

    def initUI(self):
        """
        Initialise l'interface utilisateur de l'application avec tous les widgets et layouts.
        """
      ############ INTERFACE #############

        main_layout = QtWidgets.QHBoxLayout() # Création du layout principal horizontal
        
        # Layout gauche
        layout = QtWidgets.QVBoxLayout() # Création du layout vertical à gauche

        # Label pour le chemin d'accès Excel
        self.label_excel_path = QLabel("Chemin d'accès Excel :")
        layout.addWidget(self.label_excel_path)

        # Chemin d'accès excel
        self.path_excel = QtWidgets.QLineEdit(get_config_value("SETTINGS", "excel_path"))  
        self.path_excel.setReadOnly(True)
        layout.addWidget(self.path_excel)

        # Boutons Modifier et Ouvrir excel
        button_layout_excel = QtWidgets.QHBoxLayout()
        self.modify_button_excel = QtWidgets.QPushButton('Modifier')
        self.open_button_excel = QtWidgets.QPushButton('Ouvrir')
        button_layout_excel.addWidget(self.modify_button_excel)
        button_layout_excel.addWidget(self.open_button_excel)
        self.modify_button_excel.setToolTip('Cliquez ici pour modifier le chemin')

        layout.addLayout(button_layout_excel)

        # Label pour le chemin d'accès PDF
        self.label_pdf_path = QLabel("Chemin d'accès PDF :")
        layout.addWidget(self.label_pdf_path)

        # Chemin d'accès PDF
        path_pdf = get_config_value("SETTINGS", "pdf_path")
        self.path_pdf = QtWidgets.QLineEdit(path_pdf)
        self.path_pdf.setReadOnly(True)
        layout.addWidget(self.path_pdf)

        # Boutons Modifier PDF
        button_layout_pdf = QtWidgets.QHBoxLayout()
        self.modify_button_pdf = QtWidgets.QPushButton('Modifier')
        button_layout_pdf.addWidget(self.modify_button_pdf)
        layout.addLayout(button_layout_pdf)

         # Label pour le nom PDF
        self.label_name_pdf_path = QLabel("Nom du PDF :")
        layout.addWidget(self.label_name_pdf_path)

        # Nom PDF
        name_pdf = get_config_value("SETTINGS", "name_pdf")
        self.path_namepdf = QtWidgets.QLineEdit(name_pdf)
        self.path_namepdf.setReadOnly(True)
        layout.addWidget(self.path_namepdf)

        # Boutons Modifier nom PDF
        button_layout_namepdf = QtWidgets.QHBoxLayout()
        self.modify_button_namepdf = QtWidgets.QPushButton('Modifier')
        button_layout_namepdf.addWidget(self.modify_button_namepdf)
        layout.addLayout(button_layout_namepdf)

        # Connexion Buttons avec fonctions
        self.modify_button_excel.clicked.connect(self.modify_path_excel)
        self.open_button_excel.clicked.connect(self.open_file_excel)
        self.modify_button_pdf.clicked.connect(self.modify_path_pdf)
        self.modify_button_namepdf.clicked.connect(self.modify_name)

        self.progressbar = QProgressBar(self)
        layout.addWidget(self.progressbar)

        # Création de la section Paramètres
        self.settings_group = QtWidgets.QGroupBox("Paramètres")
        settings_layout = QtWidgets.QVBoxLayout()


        # Bouton pour ouvrir la sélection des fonctions de scrapping
        self.select_scrapping_functions_button = QtWidgets.QPushButton('Sélectionner les Rates à récupérer.')
        self.select_scrapping_functions_button.clicked.connect(self.open_scrapping_selection_dialog)
        settings_layout.addWidget(self.select_scrapping_functions_button)

        # Date
        self.start_date_edit = QtWidgets.QDateEdit(datetime.now().date())
        self.end_date_edit = QtWidgets.QDateEdit(datetime.now().date())

        # Champs choix des dates
        self.start_date_edit.setCalendarPopup(True)
        self.end_date_edit.setCalendarPopup(True)

        settings_layout.addWidget(QtWidgets.QLabel("Date de début :"))
        settings_layout.addWidget(self.start_date_edit)
        settings_layout.addWidget(QtWidgets.QLabel("Date de fin :"))
        settings_layout.addWidget(self.end_date_edit)

        # Créez l'objet QCheckBox avant de l'ajouter au layout
        self.use_date_range_checkbox = QtWidgets.QCheckBox("Utiliser la plage de dates")
        self.use_date_range_checkbox.stateChanged.connect(self.toggle_date_widgets)
        settings_layout.addWidget(self.use_date_range_checkbox)

        self.start_date_edit.setEnabled(self.use_date_range_checkbox.isChecked())
        self.end_date_edit.setEnabled(self.use_date_range_checkbox.isChecked())
        

        # Checkbox pour lancer automatiquement le programme
        self.param1_checkbox = QtWidgets.QCheckBox("Lancer le script automatiquement au démarrage de l'application.")
        self.param1_checkbox.stateChanged.connect(self.saveSettings)

        # Ajout des widgets au layout des paramètres
        settings_layout.addWidget(self.param1_checkbox)

        # Définition du layout des paramètres comme layout du QGroupBox
        self.settings_group.setLayout(settings_layout)

        # Ajout du QGroupBox au layout principal
        layout.addWidget(self.settings_group)


        main_layout.addLayout(layout)

        # Layout droit (Log)
        log_layout = QtWidgets.QVBoxLayout()
        self.logger = QtWidgets.QTextEdit()
        self.logger.setReadOnly(True)
        log_layout.addWidget(self.logger)

        # Ajout du layout droit au layout principal
        main_layout.addLayout(log_layout)

        # Bouton Lancer
        self.run_button = QtWidgets.QPushButton('Lancer')
        self.run_button.setFixedSize(500, 40)
        self.run_button.setStyleSheet("""
            QPushButton {
                background-color: rgba(0, 113, 255, 255);
                color: white;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: rgba(0, 113, 255, 200);
            }
        """)
        
        layout.addWidget(self.run_button)
        self.run_button.clicked.connect(lambda: self.lancer_script(sites))

        # Définition du layout principal comme layout de la fenêtre
        self.setLayout(main_layout)

        # INIT #
        self.show() # Affichage de la fenêtre

        self.update_run_button_status(day_of_week) # Mise à jour de l'état du bouton Lancer en fonction du jour de la semaine
    
    #############  FONCTIONS ###############

    def toggle_date_widgets(self):
        # Si la case est cochée
        if self.use_date_range_checkbox.isChecked():
            # Activer les widgets de date
            self.start_date_edit.setEnabled(True)
            self.end_date_edit.setEnabled(True)
        else:
            # Désactiver les widgets de date
            self.start_date_edit.setEnabled(False)
            self.end_date_edit.setEnabled(False)

    def update_run_button_status(self, day):
        # Si c'est le weekend
        if day in ["samedi", "dimanche"]:
            self.run_button.setEnabled(False) # Désactiver le bouton
            QMessageBox.information(self, "Jour fermé.", "Jour fermé, le script ne peut être lancé.")
        #Si le chemin d'accès PDF est manquant ou invalide
        elif not self.path_pdf.text().strip():
            self.run_button.setEnabled(False) # Déseactiver le bouton
            QMessageBox.information(self, "Chemin d'accès PDF manquant.", "Veuillez renseigner un chemin d'accès PDF valide.")
        else:
            self.run_button.setEnabled(True) # Activer le bouton si tout est correct

    def saveSettings(self):
        self.config.read(config_path) # Lire le fichier de configuration
    
        # Ajouter ou mettre à jour une section et une clé spécifique
        if not self.config.has_section('SETTINGS'):
            self.config.add_section('SETTINGS')
        self.config.set('SETTINGS', 'auto_start', str(self.param1_checkbox.isChecked()))
        
        # Sauvegarder les modifications dans le fichier de configuration
        with open(config_path, 'w') as configfile:
            self.config.write(configfile)

    def modify_path_excel(self):
        # Ouvrir une boîte de dialogue pour sélectionner un fichier Excel
        file_dialog = QFileDialog()
        path = file_dialog.getOpenFileName(self, 'Sélectionner un fichier Excel', '', 'Excel Files (*.xlsx *.xls)')[0]
        
        if path: # Si un chemin est sélectionné
            self.path_excel.setText(path)
            # Lire, modifier et sauvegarder le fichier de configuration
            config.read(config_path)
            if not config.has_section('SETTINGS'):
                config.add_section('SETTINGS')
            config.set('SETTINGS', 'excel_path', path)
            with open(config_path, 'w') as configfile:
                config.write(configfile)
            self.log('Chemin modifié.') # Loguer que le chemin a été modifié
    
    def modify_path_pdf(self):
        file_dialog = QFileDialog()
        path = file_dialog.getExistingDirectory(self, 'Sélectionner un dossier')
        old_path = self.path_pdf.text()
        if path: # Si un chemin est sélectionné
            self.path_pdf.setText(path)
            # Lire, modifier et sauvegarder le fichier de configuration
            config.read(config_path)
            if not config.has_section('SETTINGS'):
                config.add_section('SETTINGS')
            config.set('SETTINGS', 'pdf_path', path)
            with open(config_path, 'w') as configfile:
                config.write(configfile)
            self.log('Chemin modifié.') # Loguer que le chemin a été modifié
            if not self.restart_app():
                self.path_pdf.setText(old_path)
        self.update_run_button_status(day_of_week)
        

    def modify_name(self):
        # Une boîte de dialogue est ouverte pour permettre à l'utilisateur d'entrer un nouveau nom
        new_name, ok = QInputDialog.getText(self, 'Modifier le nom', 'Entrez le nouveau nom:')
        old_name = self.label_name_pdf_path.text() # Stocker l'ancien nom
        
        # Si l'utilisateur clique sur OK et entre un nouveau nom
        if ok and new_name:
            self.path_namepdf.setText(new_name) # Mettre à jour le nom affiché 
            
            # Lire et modifier le fichier de configuration pour réfléter le nouveau nom
            config.read(config_path)
            if not config.has_section('SETTINGS'):
                config.add_section('SETTINGS')
            config.set('SETTINGS', 'name_pdf', new_name)
            
            # Sauvegarder les modifications dans le fichier de configuration
            with open(config_path, 'w') as configfile:
                config.write(configfile)
            
            self.log('Nom modifié.') # Loguer lorsque le nom est modifié
            
            # Redémarrer l'application pour appliquer les modifications
            if not self.restart_app():
                self.label_name_pdf_path.setText(old_name)

    def open_file_excel(self):
        # Tenter d'ouvrir le fichier Excel avec une commande système
        try:
            subprocess.run(["start", default_path_excel], shell=True, check=True)
            self.log('Fichier ouvert.')# Loguer lorsque le fichier Excel est ouvert
        except subprocess.CalledProcessError as e:
            self.log('Fichier non trouvé.') # Loguer un message d'erreur si le fichier ne s'ouvre pas

    def restart_app(self):
        # Demander à l'utilisateur s'il souhaite redémarrer l'application
        reply = QMessageBox.question(
            self, "Redémarrage requis",
            "L'application doit être redémarrée pour appliquer les changements. Voulez-vous redémarrer maintenant?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        # Si l'utilisateur choisit de redémarrer, exécuter le redémarrage
        if reply == QMessageBox.Yes:
            QApplication.quit() # Quitter l'application
            os.execl(sys.executable, sys.executable, *sys.argv) # Redémarrer l'application
        else:
            return False

    def lancer_script(self, sites):
        """
        Exécute les scripts de scrapping en fonction des paramètres sélectionnés par l'utilisateur.

        Parameters:
        sites (list): Une liste de dictionnaires contenant des informations sur les sites à scraper.
        
        Cette méthode exécute plusieurs vérifications et configurations avant de lancer les scripts de scrapping.
        Elle gère également les erreurs et les exceptions qui peuvent survenir pendant le processus.
        
        """
        
        # Vérifie si des fonctions de scapping ont été sélectionnées
        if not self.selected_scrapping_functions:
            QMessageBox.warning(self, "Attention", "Veuillez selectionner au moins une Rate.")
            return
        
        # Vérifications spécifiques sur certaines Rates avec la plage de dates
        if self.use_date_range_checkbox.isChecked():
            if 'extract_2360' in self.selected_scrapping_functions:
                QMessageBox.warning(self, "Avertissement", "La fonction 2360 ne peut pas être utilisée avec une plage de dates.")
                return
            if 'extract_2CUB' in self.selected_scrapping_functions:
                QMessageBox.warning(self, "Avertissement", "La fonction 2CUB ne peut pas être utilisée avec une plage de dates.")
                return
            
        # Initialisation des variables pour la gestion des valeurs remplacées et la progress bar
        replaced_values = {}
        replaced_value_count = 0
        self.progressbar.setMaximum(len(sites))
        self.progressbar.setValue(0)
        
        # Récupération des jours fériés selon les différentes régions
        holidays_french = get_french_holidays(yesterday.year)
        holidays_uk = get_uk_holidays(yesterday.year)

        # Récupération des dates de début et de fin si la plage de dates est cochée
        if self.use_date_range_checkbox.isChecked():
            start_date = self.start_date_edit.date().toPyDate()
            end_date = self.end_date_edit.date().toPyDate()
        else:
            start_date = end_date = datetime.today().date()

         # Gestion du fichier Excel : création ou chargement
        excel_path = default_path_excel
        if not excel_path or not os.path.exists(excel_path):
            # (Code pour créer un nouveau fichier Excel)
            excel_path = os.path.join(os.getcwd(), "metals_prices.xlsx")
            wb = Workbook()
            for site in sites:
                wb.create_sheet(site['name'])

            wb.save(excel_path)
            set_config_value("SETTINGS", "excel_path", excel_path)
            self.path_excel.setText(excel_path)
        else:
            # (Code pour charger le fichier Excel s'il existe)
            wb = load_workbook(excel_path)
        # (Code pour créer l'onglet RPA)
        rpa_sheet = wb['RPA'] if 'RPA' in wb.sheetnames else wb.create_sheet('RPA')
        # (Nettoyer l'onglet RPA)
        if rpa_sheet.max_row > 1:
            rpa_sheet.delete_rows(2, rpa_sheet.max_row-1)


        # Init de la variable pour les messages d'erreur
        txterr = ""
        # Parcourir chaque site dans la liste des sites
        for site in sites:
            data_extraction_function_name = site['func']
            
            #Si la fonction de scrapping n'est pas sélectionnée, passer au site suivant
            if data_extraction_function_name not in self.selected_scrapping_functions:
                continue
            
            try:
                # Tenter de récupérer le contenu du site
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                    'Accept-Language': 'fr,fr-FR;q=0.8,en-US;q=0.5,en;q=0.3',
                }
                response = requests.get(site['url'], headers=headers, verify=False)
                response.raise_for_status()
                
            except RequestException as e:
                # En cas d'erreur, loguer le message d'erreur et passer au suivant
                txterr = f"Erreur de connexion pour le site de {site['name']} : {e}"
                self.log(txterr)
                continue
            
            # Parser le contenu de la page avec BeautifulSoup
            soup = BeautifulSoup(response.content, "html.parser")
            
            # Vérifier si la fonction de scrapping existe
            if hasattr(scrapping, data_extraction_function_name):
                
                # Si le contenu est sous forme PDF, télécharger le PDF
                if site['src'] == 'pdf':
                    download_pdf(response, site['name_pdf'], default_path_pdf)
                
                # Obtenir la fonction de scrapping et la feuille Excel correspondante
                data_extraction_function = getattr(scrapping, data_extraction_function_name)
                sheet = wb[site["name"]]

                # Extraire les données en utilisant la fonction de scrapping
                extracted_data = data_extraction_function(soup, checkbox_state=self.use_date_range_checkbox.isChecked(), start_date=start_date, end_date=end_date)
                data = None
                
                # Mettre à jour la barre de progression
                self.progressbar.setValue(self.progressbar.value() + 1)
                
                # Si la plage de dates est cochée, écrire chaque paire de données extraites dans la feuille Excel
                if self.use_date_range_checkbox.isChecked():
                    for date_day, data in extracted_data:
                        row_number = sheet.max_row + 1
                        sheet.cell(row=row_number, column=1, value=date_day)
                        sheet.cell(row=row_number, column=2, value=data)
                        sheet.cell(row=row_number, column=3, value=site['devise'])
                        sheet.cell(row=row_number, column=4, value=site['unit'])
                        self.log(f"Valeur inscrite : {data} avec la date : {date_day} pour le site {site['name']}")

                else:
                    # Si plage de dates n'est pas cochée, écrire seulement la première paire de données extraites dans la feuille Excel
                    date_day, data = extracted_data
                    print(date_day, data)
                    row_number = sheet.max_row + 1
                    if date_day == 'date none' and data == 'value none':
                        # Récupérez la date et la valeur de la ligne précédente
                        prev_date = sheet.cell(row=row_number - 1, column=1).value
                        prev_value = sheet.cell(row=row_number - 1, column=2).value
                        print(f"prev value {prev_value}")
                        sheet.cell(row=row_number, column=1, value=prev_date)
                        sheet.cell(row=row_number, column=2, value=prev_value)
                        replaced_value_count += 1
                        replaced_values[f"Rate {site['name']}"] = f'Date: {prev_date}, Value: {prev_value}'
                    else: 
                        if jour_actuel == 0:
                            date_attendue = date_actuelle - timedelta(days=(date_actuelle.weekday() + 3))
                            
                        else:
                            date_attendue = date_actuelle - timedelta(days=1)
                        
                        # Expression régulière pour détecter le format "Semaine XX"
                        semaine_regex = r"^Semaine \d{1,2}$"
                        date_day = date_day.strip()
                        if re.match(semaine_regex, date_day):
                            print("Format 'Semaine XX' détecté, ignoré")
                            sheet.cell(row=row_number, column=1, value=date_day)
                            sheet.cell(row=row_number, column=2, value=data)
                            self.log(f"Valeur inscrite : {data} avec la date : {date_day} pour le site {site['name']}")
                        else:
                            date_day = date_day.strip()
                            try:
                                date_day = datetime.strptime(date_day, "%d/%m/%Y")
                                is_correct_date = date_day.date() == date_attendue.date() and date_day.weekday() == jour_attendu
                                print(date_day)
                            except ValueError:
                                print("Format de date non valide")
                                is_correct_date = False
                            if is_correct_date:
                                formatted_date = date_day.strftime("%d/%m/%Y")
                                print("La date correspond, écriture dans Excel")
                                print(f"Valeur inscrite : {data} avec la date : {formatted_date}")
                                sheet.cell(row=row_number, column=1, value=formatted_date)
                                sheet.cell(row=row_number, column=2, value=data)
                                self.log(f"Valeur inscrite : {data} avec la date : {formatted_date} pour le site {site['name']}")
                            else:
                                replaced_value_count += 1
                                replaced_values[f"Rate {site['name']}"] = f'Date: {date_day}, Value: {data}'
                            
                                sheet.cell(row=row_number, column=1, value=date_day)
                                sheet.cell(row=row_number, column=2, value=data)
                                self.log(f"Valeur inscrite : {data} avec la date : {date_day} pour le site {site['name']}")

                    sheet.cell(row=row_number, column=3, value=site['devise'])
                    sheet.cell(row=row_number, column=4, value=site['unit'])

                 # Écrire toutes les valeurs dans l'onglet RPA
                    extracted_data = [extracted_data]
                    print(f" Print RPA : {extracted_data} ")
                    for date_day, data in extracted_data:
                        date_str = date_day.strip()
                        if not re.match(r'Semaine \d+', date_str):
                            date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                            if isinstance(date_obj, datetime):
                                locale.setlocale(locale.LC_TIME, 'fr_FR')
                                jour_de_la_semaine = date_obj.strftime('%A')
                                date_formated = date_obj.strftime('%d/%m/%Y')
                                rpa_sheet.cell(row=1, column=1, value=date_formated)
                                rpa_sheet.cell(row=1, column=2, value=jour_de_la_semaine)
                        else:
                            pass
                         
                        rpa_row_number = rpa_sheet.max_row + 1
                        rpa_sheet.cell(row=rpa_row_number, column=1, value=site['metal'])
                        rpa_sheet.cell(row=rpa_row_number, column=2, value=site['name'])
                        
                        data_sheet = wb[site['name']]
                        if date_day == 'date none' and data == 'value none':
                            prev_date = data_sheet.cell(row=rpa_row_number - 1, column=1).value
                            prev_value = data_sheet.cell(row=rpa_row_number - 1, column=2).value

                            rpa_sheet.cell(row=rpa_row_number, column=3, value=prev_value)
                        else:
                            rpa_sheet.cell(row=rpa_row_number, column=3, value=data)
                        rpa_sheet.cell(row=rpa_row_number, column=4, value=site['devise'])
                        rpa_sheet.cell(row=rpa_row_number, column=5, value=site['unit'])
                
                self.log(txterr)
                wb.save(excel_path)
                print("Saved")
            else:
                print(f'Aucune fonction d\'extraction de données trouvées')
        replaced_message = f"{replaced_value_count} valeurs remplacées : {', '.join(f'{k}: {v}' for k, v in replaced_values.items())}"
        self.log("Script terminé.")
        QMessageBox.information(self, "Information", f"Le script a terminé l'extraction des données et la mise à jour du fichier Excel.\n{replaced_message}")
    
    def log(self, message):
        # Fonction pour log les messages
        self.logger.append(message)